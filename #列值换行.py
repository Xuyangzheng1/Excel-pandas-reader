#3
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

def wrap_text_in_all_cells_except_first_row(file_path, max_characters):
    try:
        # 打开Excel文件
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # 创建新的Font对象，设置字体为微软雅黑
        new_font = Font(name='Microsoft YaHei')

        # 创建新的Alignment对象，设置文本居中和自动换行
        new_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 获取所有行的最大列数和最大行数
        max_columns = sheet.max_column
        max_rows = sheet.max_row

        # 遍历除第一行之外的所有单元格
        for row in range(2, max_rows + 1):
            for col in range(1, max_columns + 1):
                cell = sheet.cell(row=row, column=col)

                # 设置新的字体样式和对齐样式
                cell.font = new_font
                cell.alignment = new_alignment

                # 如果单元格中的字符超过 max_characters，则将其分隔为每行不超过 max_characters 个字符的多行字符串
                if len(str(cell.value)) > max_characters:
                    split_text = [str(cell.value)[i:i+max_characters] for i in range(0, len(str(cell.value)), max_characters)]
                    cell.value = '\n'.join(split_text)

        # 保存更改后的文件
        workbook.save(file_path)
        workbook.close()

    except Exception as e:
        print(f"处理文件 {file_path} 时出现错误: {e}")

def process_xlsx_files_in_directory(directory, max_characters):
    files = os.listdir(directory)
    xlsx_files = [file for file in files if file.endswith('.xlsx')]

    if xlsx_files:
        for xlsx_file in xlsx_files:
            file_path = os.path.join(directory, xlsx_file)
            wrap_text_in_all_cells_except_first_row(file_path, max_characters)
            print(f"处理文件: {xlsx_file}")

    else:
        print("指定目录中没有XLSX文件")

# 指定每个单元格字符的最大数量
max_characters = 6

# 指定目录的路径
directory_path = r"K:\医保局2024\职工大额格式\excel数据职工大额test"

# 调用函数处理XLSX文件
process_xlsx_files_in_directory(directory_path, max_characters)
