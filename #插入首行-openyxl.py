#4
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

def add_blank_row_and_merge_cells_above_header(directory):
    # 获取目录下所有xlsx文件
    xlsx_files = [file for file in os.listdir(directory) if file.endswith('.xlsx')]

    # 遍历每个xlsx文件
    for xlsx_file in xlsx_files:
        file_path = os.path.join(directory, xlsx_file)

        # 读取xlsx文件
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # 在列名行之上插入一行完全空白的行
        sheet.insert_rows(1)
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=1, column=col).value = ''

        # 合并新插入行的所有单元格
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet.max_column)

        # 添加值，当前读取的xlsx文件名（去掉扩展名）和医保系统结算信息
        file_name_without_extension = os.path.splitext(xlsx_file)[0]
        sheet.cell(row=1, column=1).value = f"{file_name_without_extension} （医保系统结算信息）"

        # 设置字体为微软雅黑，大小为28号，居中
        font = Font(name='微软雅黑', size=28)
        alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=1, column=1).font = font
        sheet.cell(row=1, column=1).alignment = alignment

        # 设置整个表格的边框为实线
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = Border(top=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))

        # 保存修改后的数据
        workbook.save(file_path)

if __name__ == "__main__":
    # 指定目录路径
    target_directory = r"K:\医保局2024\职工大额格式\excel数据职工大额test"

    # 调用函数
    add_blank_row_and_merge_cells_above_header(target_directory)
