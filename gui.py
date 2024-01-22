#医保信息处理-源码-打包
#医保信息处理-源码
#1==========================xls转xlsx
import os
import glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import tkinter as tk
from tkinter import filedialog
from tkinter import simpledialog
from tkinter import messagebox
class ExcelProcessorGUI:
    def __init__(self, master):
        self.master = master
        master.title("Excel Processor GUI")

        # Create GUI elements
        self.max_characters_header_label = tk.Label(master, text="Max Characters for Header:")
        self.max_characters_label = tk.Label(master, text="Max Characters for Cells:")
        self.directory_label = tk.Label(master, text="Target Directory:")
        
        self.max_characters_header_entry = tk.Entry(master)
        self.max_characters_entry = tk.Entry(master)
        self.directory_entry = tk.Entry(master, state='disabled')

        self.browse_button = tk.Button(master, text="Browse", command=self.browse_directory)
        self.process_button = tk.Button(master, text="Process Files", command=self.process_files)

        # Set default values
        self.max_characters_header_entry.insert(0, "5")
        self.max_characters_entry.insert(0, "7")

        # Grid layout
        self.max_characters_header_label.grid(row=0, column=0, sticky='e')
        self.max_characters_header_entry.grid(row=0, column=1)
        self.max_characters_label.grid(row=1, column=0, sticky='e')
        self.max_characters_entry.grid(row=1, column=1)
        self.directory_label.grid(row=2, column=0, sticky='e')
        self.directory_entry.grid(row=2, column=1)
        self.browse_button.grid(row=2, column=2)
        self.process_button.grid(row=3, column=1)

    def browse_directory(self):
        directory = filedialog.askdirectory()
        self.directory_entry.config(state='normal')
        self.directory_entry.delete(0, tk.END)
        self.directory_entry.insert(0, directory)
        self.directory_entry.config(state='disabled')

    def process_files(self):
        try:
            max_characters_header = int(self.max_characters_header_entry.get())
            max_characters = int(self.max_characters_entry.get())
            target_directory = self.directory_entry.get()

            # Call your processing methods here with the user-selected values
            converter = XlsToXlsxConverter(target_directory)
            excel_wrapper = ExcelTextWrapper(target_directory, max_characters_header)
            excel_processor = ExcelProcessor(max_characters)
            excel_processor_header = ExcelProcessor_header(target_directory)

            converter.convert_xls_to_xlsx_in_directory()
            excel_wrapper.process_xlsx_files_in_directory()
            excel_processor.process_xlsx_files_in_directory(target_directory)
            excel_processor_header.process_excel_files()

            messagebox.showinfo("Success", "Files processed successfully!")

        except ValueError:
            messagebox.showerror("Error", "Invalid input. Please enter valid integer values for max characters.")

class XlsToXlsxConverter:
    def __init__(self, target_directory):
        self.target_directory = target_directory

    def xls_to_xlsx(self, input_path, output_path):
        xls_data = pd.read_excel(input_path, None)
        with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
            for sheet_name, df in xls_data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    def convert_xls_to_xlsx_in_directory(self):
        xls_files = glob.glob(os.path.join(self.target_directory, '*.xls'))
        xlsx_files = glob.glob(os.path.join(self.target_directory, '*.xlsx'))
        all_files = xls_files + xlsx_files
        print(all_files)

        for xls_file in xls_files:
            xls_path = os.path.abspath(xls_file)
            xlsx_path = os.path.splitext(xls_path)[0] + '.xlsx'
            self.xls_to_xlsx(xls_path, xlsx_path)
            os.remove(xls_path)
#2========================================================
           


class ExcelTextWrapper:
    def __init__(self, directory, max_characters):
        self.directory = directory
        self.max_characters = max_characters

    def wrap_text_in_first_row(self, file_path):
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active

            new_font = Font(name='Microsoft YaHei')
            new_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            first_row_cells = sheet[1]

            for cell in first_row_cells:
                cell.font = new_font
                cell.alignment = new_alignment

                if len(str(cell.value)) > self.max_characters:
                    split_text = [str(cell.value)[i:i+self.max_characters] for i in range(0, len(str(cell.value)), self.max_characters)]
                    cell.value = '\n'.join(split_text)

            workbook.save(file_path)
            workbook.close()

        except Exception as e:
            print(f"处理文件 {file_path} 时出现错误: {e}")

    def process_xlsx_files_in_directory(self):
        files = os.listdir(self.directory)
        xlsx_files = [file for file in files if file.endswith('.xlsx')]

        if xlsx_files:
            for xlsx_file in xlsx_files:
                file_path = os.path.join(self.directory, xlsx_file)
                self.wrap_text_in_first_row(file_path)
                print(f"处理文件: {xlsx_file}")

        else:
            print("指定目录中没有XLSX文件")

# if __name__ == "__main__":
#     # 指定每个单元格字符的最大数量
#     max_characters = 4

#     # 指定目录的路径
#     directory_path = r"K:\医保局2024\职工大额格式\excel数据职工大额test"

#     # 创建ExcelTextWrapper类的实例
#     excel_wrapper = ExcelTextWrapper(directory_path, max_characters)

#     # 调用方法处理XLSX文件
#     excel_wrapper.process_xlsx_files_in_directory()
#3=============================================


class ExcelProcessor:
    def __init__(self, max_characters=6):
        self.max_characters = max_characters
        self.new_font = Font(name='Microsoft YaHei')
        self.new_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def wrap_text_in_all_cells_except_first_row(self, file_path):
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active

            max_columns = sheet.max_column
            max_rows = sheet.max_row

            for row in range(2, max_rows + 1):
                for col in range(1, max_columns + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell.font = self.new_font
                    cell.alignment = self.new_alignment

                    if len(str(cell.value)) > self.max_characters:
                        split_text = [str(cell.value)[i:i + self.max_characters] for i in range(0, len(str(cell.value)), self.max_characters)]
                        cell.value = '\n'.join(split_text)

            workbook.save(file_path)
            workbook.close()

        except Exception as e:
            print(f"Error processing file {file_path}: {e}")

    def process_xlsx_files_in_directory(self, directory):
        files = os.listdir(directory)
        xlsx_files = [file for file in files if file.endswith('.xlsx')]

        if xlsx_files:
            for xlsx_file in xlsx_files:
                file_path = os.path.join(directory, xlsx_file)
                self.wrap_text_in_all_cells_except_first_row(file_path)
                print(f"Processed file: {xlsx_file}")
        else:
            print("No XLSX files in the specified directory")

# Example Usage:
# max_characters = 6
# directory_path = r"K:\医保局2024\职工大额格式\excel数据职工大额test"

# excel_processor = ExcelProcessor(max_characters)
# excel_processor.process_xlsx_files_in_directory(directory_path)

#4================================================添加标题完善表格


class ExcelProcessor_header:
    def __init__(self, directory):
        self.directory = directory

    def _process_excel_file(self, file_path):
        # 读取xlsx文件
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # 在列名行之上插入一行完全空白的行
        sheet.insert_rows(1)
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=1, column=col).value = ''

        # 合并新插入行的所有单元格
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sheet.max_column)

        # 添加值，当前读取的xlsx文件名（去掉扩展名）和医保系统结算信息
        file_name_without_extension = os.path.splitext(os.path.basename(file_path))[0]
        sheet.cell(row=1, column=1).value = f"{file_name_without_extension} （医保系统结算信息）"

        # 设置字体为微软雅黑，大小为28号，居中
        font = Font(name='微软雅黑', size=28)
        alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=1, column=1).font = font
        sheet.cell(row=1, column=1).alignment = alignment

        # 设置整个表格的边框为实线
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = Border(top=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'))

        # 保存修改后的数据
        workbook.save(file_path)

    def process_excel_files(self):
        # 获取目录下所有xlsx文件
        xlsx_files = [file for file in os.listdir(self.directory) if file.endswith('.xlsx')]

        # 遍历每个xlsx文件
        for xlsx_file in xlsx_files:
            file_path = os.path.join(self.directory, xlsx_file)
            self._process_excel_file(file_path)






if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelProcessorGUI(root)
    root.mainloop()

    # max_characters_header = 5
    # max_characters = 7


    # target_directory = r'K:\医保局2024\职工大额格式\excel数据职工大额test'
    
    # # 创建XlsToXlsxConverter类的实例
    # converter = XlsToXlsxConverter(target_directory)
    # excel_wrapper = ExcelTextWrapper(target_directory, max_characters_header)
    # excel_processor = ExcelProcessor(max_characters)
    # excel_processor_header = ExcelProcessor_header(target_directory)
    
    # # 调用方法，在指定目录中将xls文件转换为xlsx格式
    # converter.convert_xls_to_xlsx_in_directory()
    # print(111111111111111111111)
    # excel_wrapper.process_xlsx_files_in_directory()
    # print(222222222222222222222)
    # excel_processor.process_xlsx_files_in_directory(target_directory)
    # print(333333333333333333333)
    # excel_processor_header.process_excel_files()
    # print(4444444444444444444444)
